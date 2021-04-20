import openpyxl


def getRowCount(file):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    return (sheet.max_row)


def getColumnCount(file):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    return (sheet.max_column)


def readData(file, rowum, column_text):
    workbook = openpyxl.load_workbook(file, data_only=True)
    sheet = workbook.active
    column_count = sheet.max_column
    columnno = 0
    for i in range(1, column_count + 1):
        heading = sheet.cell(row=1, column=i).value.lower()
        if heading == column_text.lower():
            columnno = i
            break
    try:
        return sheet.cell(row=rowum, column=columnno).value
    except:
        return 'Null'


def writeData(file, rownum, column_text, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    column_count = sheet.max_column
    columnno = 0
    for i in range(1, column_count + 1):
        heading = sheet.cell(row=1, column=i).value.lower()
        if heading == column_text.lower():
            columnno = i
            break
    sheet.cell(row=rownum, column=columnno).value = data
    workbook.save(file)
